from openpyxl import load_workbook
#from openpyxl.cell import get_column_letter, column_index_from_string
from openpyxl.utils import get_column_letter


def get_data(path):
    #path = './Static/13区.xlsx'
    wb = load_workbook(path)

    sheet = wb.get_sheet_by_name('Sheet1')

    matrix = max(sheet.max_column, sheet.max_row)
    a = []
    count = 0
    for i in range(1,sheet.max_row+1):
        temp = []
        for j in range(1, sheet.max_column+1):
            letter = get_column_letter(j)
            value = sheet[letter+str(i)].value
            if value:
                if value == 1:
                    temp.append('sell')
                    count += 1
                elif value == 2:
                    temp.append('protect')
                elif value == 3:
                    temp.append('cover')
            else:
                temp.append(0)
        a.append(temp)
    return a, count

